from flask import Flask, request, render_template, redirect, session, flash, url_for,send_file, jsonify, abort, Response, make_response, send_file
from models import db, User, EmployerDetails, Jobs, Application, SPES, SPESApplication, PESO, Task
from flask_mail import Mail, Message
import random
import base64, io
from datetime import datetime
from flask_ngrok import run_with_ngrok
import webbrowser
from io import BytesIO
from sqlalchemy import not_, or_
from flask_login import current_user
from io import BytesIO
from pdfcrowd import HtmlToPdfClient
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import openpyxl
from io import BytesIO
import os

app = Flask(__name__)

app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get("DATABASE_URL")
# postgres://pilahunt_7hbx_user:dxvqEXIt0XpCePejZdo7QjScRfSPlpoA@dpg-cper33f79t8c73bcor90-a.singapore-postgres.render.com/pilahunt_7hbx

app.secret_key = 'secret_key'

app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'pilahuntpilapeso@gmail.com'
app.config['MAIL_PASSWORD'] = 'wobgjtkqpekgjcmd'
app.config['MAIL_DEFAULT_SENDER'] = 'pilahuntpilapeso@gmail.com'

mail = Mail(app)


db.init_app(app)
with app.app_context():
    db.create_all()

####EMPLOYERS/EMPLOYEESSSS#####
@app.route('/')
def landingpage():
    return render_template('landingpage.html')


@app.route('/applicant_login', methods=['GET', 'POST'])
def applicant_login():
    if request.method == 'POST':
        if 'login' in request.form:
            email = request.form['email']
            password = request.form['password']
            user = User.query.filter_by(email=email).first()
            employer = EmployerDetails.query.filter_by(email_add=email).first()

            if user and user.check_password(password):
                session['email'] = user.email
                if user.user_type == 'applicant':
                    return redirect('/home')  
                elif user.user_type == 'employer':
                    if employer.verification_status in ['not verified', 'waiting verification', 'Under Review']:
                        error = "Your account is under verification."
                        return render_template('applicant-login.html', error=error)
                    else:
                        return redirect('/employer_home')
            else:
                return render_template('applicant-login.html', error='Invalid email or password')
        
        elif 'signup' in request.form:
            name = request.form['name']
            email = request.form['email']
            firstname = request.form['firstname']
            lastname = request.form['lastname']
            password = request.form['password']
            confirm_password = request.form['confirm_password']
            user_type = request.form['user_type']

            existing_user = User.query.filter_by(email=email).first()
            if existing_user:
                error = "Email already exists. Please choose a different email."
                return render_template('applicant-login.html', error=error)


            if password != confirm_password:
                error = "Passwords do not match!"
                return render_template('applicant-login.html', error=error)

            verification_code = ''.join(random.choices('0123456789', k=5))

            send_verification_code_email(email, verification_code)

            session['signup_details'] = {
                'name': name,
                'email': email,
                'password': password,
                'user_type': user_type,
                'firstname': firstname,
                'lastname': lastname
                
            }
            session['verification_code'] = verification_code

            flash('Verification code has been sent to your email address. Please check your inbox.', 'success')
            return redirect('/verify')
            
    return render_template('applicant-login.html')

@app.route('/login-signup', methods=['GET', 'POST'])
def login_signup():
    if request.method == 'POST':
        if 'login' in request.form:
            email = request.form['email']
            password = request.form['password']
            user = User.query.filter_by(email=email).first()
            employer = EmployerDetails.query.filter_by(email_add=email).first()

            if user and user.check_password(password):
                session['email'] = user.email
                if user.user_type == 'applicant':
                    return redirect('/home')
                elif user.user_type == 'employer':
                    if employer.verification_status in ['not verified', 'waiting verification', 'Under Review']:
                        error = "Your account is under verification."
                        return render_template('login-signup.html', error=error)
                    else:
                        return redirect('/employer_home')
            else:
                error = 'Invalid email or password'
                return render_template('login-signup.html', error=error)

        elif 'signup' in request.form:
            name = request.form['name']
            email = request.form['email']
            password = request.form['password']
            confirm_password = request.form['confirm_password']
            user_type = request.form['user_type']
            tin_id = request.form['tin_id']
            company_name = request.form['company_name']
            company_address = request.form['company_address']
            trade_name = request.form['trade_name']
            employer_type = request.form['employer_type']
            business_type = request.form['business_type']
            contact_person = request.form['contact_person']
            phone_number = request.form['phone_no']
            position = request.form['position']

            existing_user = User.query.filter_by(email=email).first()
            if existing_user:
                error = "Email already exists. Please choose a different email."
                return render_template('login-signup.html', error=error)

            if password != confirm_password:
                error = "Passwords do not match!"
                return render_template('login-signup.html', error=error)

            verification_code = ''.join(random.choices('0123456789', k=5))

            send_verification_code_email(email, verification_code)

            session['signup_details'] = {
                'name': name,
                'email': email,
                'password': password,
                'user_type': user_type,
                'tin_id': tin_id,
                'company_name': company_name,
                'company_address': company_address,
                'trade_name': trade_name,
                'employer_type': employer_type,
                'business_type': business_type,
                'contact_person': contact_person,
                'phone_no': phone_number,
                'position': position       
            }
            session['verification_code'] = verification_code

            flash('Verification code has been sent to your email address. Please check your inbox.', 'success')
            return redirect('/verify')

    return render_template('login-signup.html')


def send_verification_code_email(email, verification_code):
    subject = "Verification Code"
    body = f"""Dear User,

Thank you for using our service. To verify your email address, please use the following verification code:

Verification Code: {verification_code}

Please enter this code in the appropriate field to complete the verification process.

If you did not request this verification code, please disregard this email.

Thank you,
PilaHunt Team"""

    msg = Message(subject, recipients=[email])
    msg.body = body
    mail.send(msg)

@app.route('/verify', methods=['GET', 'POST'])
def verify():
    if request.method == 'POST':
        entered_code = request.form['verification_code']
        if entered_code == session.get('verification_code'):
            signup_details = session.get('signup_details')
            new_user = User(
                name=signup_details['name'], 
                email=signup_details['email'],
                firstname=signup_details['firstname'],
                lastname=signup_details['lastname'], 
                password=signup_details['password'],
                user_type=signup_details['user_type'],
                user_status='active',
            )
            db.session.add(new_user)
            db.session.commit()

            if signup_details['user_type'] == 'employer':
                employer_details = EmployerDetails(
                    user_id=new_user.id,
                    verification_status="not verified",
                    tin_id=signup_details['tin_id'],
                    company_name=signup_details['company_name'],
                    company_address=signup_details['company_address'],
                    trade_name=signup_details['trade_name'],
                    employer_type=signup_details['employer_type'],
                    business_type=signup_details['business_type'],
                    contact_person=signup_details['contact_person'],
                    phone_no=signup_details['phone_no'],
                    position=signup_details['position'],
                    email_add=signup_details['email']
                )
                db.session.add(employer_details)
                db.session.commit()

            session.pop('signup_details')
            session.pop('verification_code')
            flash('Signup successful! You can now log in.', 'success')
            return redirect('/applicant_login')
        else:
            # If verification code doesn't match, show error message in the form
            error = 'Invalid verification code. Please try again.'
            return render_template('verify.html', error=error)

    return render_template('verify.html')

@app.route('/home')
def home():
    if session.get('email'):
        user = User.query.filter_by(email=session['email']).first()
        return render_template('home.html', user=user)
    return redirect('/login-signup')


@app.route('/employer_home')
def employer_home():
    if session.get('email'):
        user = User.query.filter_by(email=session['email']).first()
        return render_template('employer_home.html', user=user)
    return redirect('/login-signup')

@app.route('/view_all_jobs') 
def view_all_jobs():
    if session.get('email'):
        user = User.query.filter_by(email=session['email']).first()
        jobs = Jobs.query.all()
        return render_template('view_all_jobs.html', user=user, jobs=jobs)
    return redirect('/login-signup')

from sqlalchemy import or_, func

@app.route('/job_listing')
def job_listing():
    if session.get('email'):
        user = User.query.filter_by(email=session['email']).first()
        job_count = Jobs.query.filter(Jobs.job_status != 'inactive').count()
        skills = user.skills.split(',') if user.skills else []
        page = request.args.get('page', 1, type=int)
        per_page = 5  

        # Filter jobs based on user skills
        filtered_jobs = Jobs.query.filter(
            (Jobs.job_status != 'inactive') &
            (or_(
                *[
                    func.lower(Jobs.job_requirements).like(f"%{skill.strip()}%")
                    for skill in skills
                ]
            ))
        )
        
        # Paginate the filtered jobs
        jobs = filtered_jobs.paginate(page=page, per_page=per_page)
        
        return render_template('job_listing.html', user=user, jobs=jobs, job_count=job_count)
    return redirect('/login-signup')


@app.route('/job_applied_list')
def job_applied_list():
    if session.get('email'):
        user = User.query.filter_by(email=session['email']).first()
        if user:
            # Query applications for the specific user
            applications = Application.query.filter_by(user_id=user.id).all()
            return render_template('job_applied_list.html', user=user, applications=applications)
        else:
            flash('User not found. Please login.', 'error')
            return redirect(url_for('login-signup'))
    else:
        flash('Please login to view job applications.', 'error')
        return redirect(url_for('login-signup'))
    
@app.route('/view_application/<int:application_id>')
def view_application(application_id):
    application = Application.query.get(application_id)
    if application:
        return render_template('view_application.html', application=application)
    else:
        flash('Job application not found.', 'error')
        return redirect(url_for('job_applied_list'))
    
@app.route('/cancel_job_application', methods=['POST'])
def cancel_job_application():
    if request.method == 'POST':
        application_id = request.form.get('application_id')
        application = Application.query.get(application_id)

        if application:
            if application.application_status == 'Under review':
                flash('You cannot cancel the application while it is under review.', 'error')
            else:
                db.session.delete(application)
                db.session.commit()
                flash('Job application cancelled successfully!', 'success')
        else:
            flash('Application not found.', 'error')

        return redirect(url_for('job_applied_list')) 


@app.route('/setup_account', methods=['POST'])
def setup_account():
    if request.method == 'POST':
        if session.get('email'):
            tin_id = request.form['tin_id']
            company_name = request.form['company_name']
            company_address = request.form['company_address']
            trade_name = request.form['trade_name']
            employer_type = request.form['employer_type']
            business_type = request.form['business_type']
            contact_person = request.form['contact_person']
            position = request.form['position']
            phone_number = request.form['phone_no']
            email_address = request.form['email_add']

            user = User.query.filter_by(email=session['email']).first()

            if user:
                existing_details = EmployerDetails.query.filter_by(user_id=user.id).first()

                if existing_details:
                    existing_details.tin_id = tin_id
                    existing_details.company_name = company_name
                    existing_details.company_address = company_address
                    existing_details.trade_name = trade_name
                    existing_details.employer_type = employer_type
                    existing_details.business_type = business_type
                    existing_details.contact_person = contact_person
                    existing_details.position = position
                    existing_details.phone_no = phone_number
                    existing_details.email_add = email_address
                    employer_details = existing_details
                else:
                    employer_details = EmployerDetails(
                        tin_id=tin_id,
                        company_name=company_name,
                        company_address=company_address,
                        trade_name=trade_name,
                        employer_type=employer_type,
                        business_type=business_type,
                        contact_person=contact_person,
                        position=position,
                        phone_no=phone_number,
                        email_add=email_address,
                        user_id=user.id  
                    )
                    db.session.add(employer_details)

                db.session.commit()

                jobs_with_employer_id = Jobs.query.filter_by(employer_id=employer_details.company_id).all()
                for job in jobs_with_employer_id:
                    job.company_name = company_name

                return redirect(url_for('employer_setup'))
            else:
                return redirect('/login-signup')
        else:
            return redirect('/login-signup')
    return redirect(url_for('employer_profile'))


@app.route('/post_job', methods=['GET', 'POST'])
def post_job():
    if request.method == 'POST':
        job_title = request.form['job_title']
        job_address = request.form['job_address']
        job_description = request.form['job_description']
        job_requirements = request.form['job_requirements']
        job_benefits = request.form['job_benefits']
        salary_range_min = float(request.form['salary_range_min'])
        salary_range_max = float(request.form['salary_range_max'])
        job_type = request.form['job_type']
        job_status = request.form['status']

        contact_person = request.form['contact_person']
        phone_no = request.form['phone_no']
        email_add = request.form['email_add']

        if session.get('email'):
            user = User.query.filter_by(email=session['email']).first()

            if user:
                employer = EmployerDetails.query.filter_by(user_id=user.id).first()  

                if employer and (employer.verification_status == "not verified" or employer.verification_status == "waiting verification" or employer.verification_status == "Under Review"):
                    flash('You need to fully verify your account before posting a job. Please Complete Account Verification.', 'error')
                    return redirect('/employer_setup')  

                new_job = Jobs(
                    job_title=job_title,
                    job_address=job_address,
                    job_description=job_description,
                    job_requirements=job_requirements,
                    salary_range_min=salary_range_min,
                    salary_range_max=salary_range_max,
                    job_benefits=job_benefits,
                    job_type=job_type,
                    job_status=job_status,
                    employer_id=employer.company_id,  # Ensure you are using the correct foreign key field
                    company_name=employer.company_name,
                    contact_person=contact_person,
                    phone_no=phone_no,
                    email_add=email_add
                )

                db.session.add(new_job)
                db.session.commit()

                flash('Job posted successfully!', 'success')
                return render_template('post_job.html', employer=employer)
            else:
                flash('User not found. Please login or sign up to post a job.', 'error')
                return redirect('/login-signup')
        else:
            flash('Please login or sign up to post a job.', 'error')
            return redirect('/login-signup')
    else:
        if session.get('email'):
            user = User.query.filter_by(email=session['email']).first()

            if user:
                employer = EmployerDetails.query.filter_by(user_id=user.id).first()
                return render_template('post_job.html', employer=employer)
            else:
                flash('User not found. Please login or sign up to post a job.', 'error')
                return redirect('/login-signup')
        else:
            flash('Please login or sign up to post a job.', 'error')
            return redirect('/login-signup')

@app.route('/view_posted_jobs')
def view_posted_jobs():
    if session.get('email'):
        user = User.query.filter_by(email=session['email']).first()
        if user and user.employer_details:
            employer_id = user.employer_details.company_id
            page = request.args.get('page', 1, type=int)
            per_page = 5
            jobs = Jobs.query.filter_by(employer_id=employer_id).paginate(page=page, per_page=per_page)
            return render_template('view_posted_jobs.html', user=user, jobs=jobs)
    return redirect('/login-signup')



        
@app.route('/posted_job_details/<int:job_id>')
def posted_job_details(job_id):
    if 'email' in session:
        user = User.query.filter_by(email=session['email']).first()
        if user:
            job = Jobs.query.get(job_id)
            if job:
                if user.user_type == 'employer':
                    employer = user.employer_details
                    if employer:
                        applications = Application.query.filter_by(job_id=job_id).all()
                        print(f"Employer details found: {employer}")  # Debugging print
                        return render_template('posted_job_details.html', user=user, job=job, employer=employer, applications=applications)
                    else:
                        print(f"No employer details found for user {user.id}")  # Debugging print
                        return "Employer details not found"
                else:
                    return "Only employers can view job details"
            else:
                return "Job not found"
        else:
            return "User not found"
    return redirect('/login-signup')

        
@app.route('/job_details/<int:job_id>')
def job_details(job_id):
    if session.get('email'):
        user = User.query.filter_by(email=session['email']).first()
        if not user:
            return redirect('/login-signup')

        employer = EmployerDetails.query.filter_by(user_id=user.id).first()  
        job = Jobs.query.get(job_id)
        if job:
            applications = Application.query.filter_by(job_id=job_id).all()
            return render_template('job_details.html', user=user, job=job, employer=employer, applications=applications)
        else:
            return render_template('error.html', message='Job not found')
    return redirect('/login-signup')

@app.route('/edit_job', methods=['GET', 'POST'])
def edit_job():
    if request.method == 'GET':
        job_id = request.args.get('job_id')
        if job_id:
            job = Jobs.query.get(job_id)
            if job:
                employer = EmployerDetails.query.get(job.employer_id)
                return render_template('edit_job.html', job=job, employer=employer)
            else:
                flash('Job not found.', 'error')
                return redirect(url_for('employer_home'))  
        else:
            flash('Job ID is missing.', 'error')
            return redirect(url_for('employer_home'))  
    elif request.method == 'POST':
        job_id = request.form.get('job_id')
        job = Jobs.query.get(job_id)
        if job:
            job.job_title = request.form['job_title']
            job.job_address = request.form['job_address']
            job.job_description = request.form['job_description']
            job.job_requirements = request.form['job_requirements']
            job.job_benefits = request.form['job_benefits']
            job.salary_range_min = float(request.form['salary_range_min'])
            job.salary_range_max = float(request.form['salary_range_max'])
            job.job_type = request.form['job_type']
            job.job_status = request.form['status']

            db.session.commit()
            flash('Job updated successfully!', 'success')
            return redirect(url_for('employer_home')) 
        else:
            flash('Job not found.', 'error')
            return redirect(url_for('employer_home'))  

@app.route('/apply_peso', methods=['POST'])
def apply_peso():
    if 'fileInput' not in request.files:
        return jsonify({'error': 'No file part'}), 400

    file = request.files['fileInput']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400

    if file and session.get('email'):
        user = User.query.filter_by(email=session['email']).first()
        if user is None:
            return jsonify({'error': 'User not found'}), 404

        # Save file to database and update work status
        user.workStatus = 'Pending'
        user.jobform = file.read()
        db.session.commit()

        # Redirect to user profile after successful upload
        return redirect(url_for('user_profile'))

    return jsonify({'error': 'Unauthorized'}), 401


@app.route('/apply', methods=['POST'])
def apply():
    if request.method == 'POST' and session.get('email'):
        user = User.query.filter_by(email=session['email']).first()
        if user:
            file = user.jobform
            company_applied = request.form['company_applied']
            position_applied = request.form['position_applied']
            company_id = request.form['company_id']
            job_id = request.form['job_id']

            existing_application = Application.query.filter_by(user_id=user.id, job_id=job_id).first()
            if existing_application:
                flash('You have already applied for this job please wait for the results of your application process.', 'error')
                return redirect(url_for('home'))

            if file:
                
                new_application = Application(
                    user_id=user.id, 
                    requirements=user.jobform,  
                    application_status='Pending',
                    application_date=datetime.now(),
                    company_id=company_id,
                    company_applied=company_applied,
                    position_applied=position_applied,
                    job_id=job_id 
                )
                db.session.add(new_application)
                db.session.commit()
                flash('Application submitted successfully!', 'success')  
                return redirect(url_for('home'))
            else:
                flash('No file uploaded!', 'error')  
                return redirect(url_for('home'))
        else:
            flash('User not found. Please login.', 'error')  
            return redirect(url_for('login-signup'))
    else:
        flash('Please login to submit an application.', 'error')  
        return redirect(url_for('login-signup'))



@app.route('/application_list')
def application_list():
    if session.get('email'):
        user = User.query.filter_by(email=session['email']).first()
        if user:
            employer = EmployerDetails.query.filter_by(user_id=user.id).first()
            if employer:
                applications = Application.query.filter_by(company_id=employer.company_id).all()
                return render_template('application_list.html', user=user, employer=employer, applications=applications)
            else:
                return render_template('error.html', message='Employer details not found')
        else:
            return render_template('error.html', message='User not found')
    return redirect('/login-signup')

        

@app.route('/applicant_details/<int:application_id>')
def applicant_details(application_id):
    if session.get('email'):
        user = User.query.filter_by(email=session['email']).first()
        employer = EmployerDetails.query.filter_by(user_id=user.id).first()  
        application = Application.query.get(application_id)
        jobs = Jobs.query.all()
        if application:
            return render_template('applicant_details.html', user=user, application=application, employer=employer, jobs=jobs)
        else:
            return render_template('error.html', message='Application not found')
    return redirect('/login-signup')

@app.route('/view_pdf/<int:application_id>')
def view_pdf(application_id):
    application = Application.query.get(application_id)
    if application:
        pdf_bytes = application.requirements
        return send_file(BytesIO(pdf_bytes), mimetype='application/pdf')
    else:
        return 'Application not found', 404
    
@app.route('/view_pdf_page/<int:application_id>')
def view_pdf_page(application_id):
    application = Application.query.get(application_id)
    if application:
        return render_template('view_pdf.html', application=application)
    else:
        return 'Application not found', 404
    
@app.route('/view_pdf_employee/<int:application_id>')
def view_pdf_employee(application_id):
    application = Application.query.get(application_id)
    if application:
        pdf_bytes = application.requirements
        return send_file(BytesIO(pdf_bytes), mimetype='application/pdf')
    else:
        return 'Application not found', 404
    
@app.route('/view_pdf_page_employee/<int:application_id>')
def view_pdf_page_employee(application_id):
    application = Application.query.get(application_id)
    if application:
        return render_template('view_pdf_page_employee.html', application=application)
    else:
        return 'Application not found', 404
    

    

@app.route('/employer_setup')
def employer_setup():
    if session.get('email'):
        user = User.query.filter_by(email=session['email']).first()
        employer = EmployerDetails.query.filter_by(user_id=user.id).first() 
        return render_template('employer_setup.html', user=user,employer=employer)
    return redirect('/login-signup')

@app.route('/change_logo', methods=['POST'])
def change_logo():
    if request.method == 'POST':
        if session.get('email'):
            user = User.query.filter_by(email=session['email']).first()
            if user:
                if 'company_logo' in request.files:
                    company_logo = request.files['company_logo']
                    if company_logo.filename != '':
                        image_data = company_logo.read()
                        base64_image = base64.b64encode(image_data).decode('utf-8')
                        employer_details = EmployerDetails.query.filter_by(user_id=user.id).first()
                        employer_details.company_logo = base64_image
                        db.session.commit()
                return redirect(url_for('employer_setup'))
            else:
                return redirect('/login-signup')
        else:
            return redirect('/login-signup')
    return redirect(url_for('employer_setup'))
 
@app.route('/logout')
def logout():
    if 'email' in session:
        session.pop('email', None)  
        if 'user_type' in session and session['user_type'] == 'employer':
            return redirect('/employer_home') 
        elif 'user_type' in session and session['user_type'] == 'applicant':
            return redirect('/home') 
        else:
            return redirect('/applicant_login') 
    elif 'spes_email' in session:
        session.pop('spes_email', None)  
        return redirect('/spes_home')
    else:
        return redirect('/applicant_login') 




def send_application_status_email(recipient_email, application_id, new_status, job_title, company_applied):
    subject = f'Application Status Update - Application ID {application_id}'
    
    body = f"The status of your application for '{job_title}' (Application ID {application_id}) at {company_applied} has been updated to: {new_status}."

    message = Message(subject, recipients=[recipient_email], body=body)
    mail.send(message)

@app.route('/update_application_status', methods=['POST'])
def update_application_status():
    if request.method == 'POST':
        application_id = request.form.get('application_id')
        status = request.form.get('status')

        application = Application.query.get(application_id)

        if application:
            application.application_status = status
            db.session.commit()

            send_application_status_email(application.user.email, application_id, status, application.job.job_title, application.company_applied)
            flash('Application status successfully updated', 'success')
        else:
            return "Application not found", 404

    return redirect(url_for('applicant_details', application_id=application_id))

#####EMPLOYEE####

def is_profile_complete(user, employer):
    user_attributes = [
        user.firstname, user.middlename, user.lastname, user.suffix, user.address, user.gender,
        user.phone, user.civil_status, user.height, user.landline, user.houseNoStreet,
        user.brgy, user.city, user.province, user.religion, user.tinID, user.gsisNO,
        user.pagibigNo, user.philNo, user.birthday  # Adjust according to what's available in User
    ]
    return all(attr is not None and attr != '' for attr in user_attributes)

@app.route('/user_profile')
def user_profile():
    if session.get('email'):
        user = User.query.filter_by(email=session['email']).first()
        employer = EmployerDetails.query.filter_by(user_id=user.id).first() 
        if user:  # Ensure both user and employer records are retrieved
            profile_complete = is_profile_complete(user, employer)
            return render_template('user_profile.html', user=user, employer=employer, profile_complete=profile_complete)
        else:
            return redirect('/login-signup')  # Redirect if no valid user or employer found
    return redirect('/login-signup')




@app.route('/change_profile', methods=['POST'])
def change_profile():
    if request.method == 'POST':
        if session.get('email'):
            user = User.query.filter_by(email=session['email']).first()
            if user:
                if 'profile_pic' in request.files:
                    profile_pic = request.files['profile_pic']
                    if profile_pic.filename != '':
                        image_data = profile_pic.read()
                        base64_image = base64.b64encode(image_data).decode('utf-8')
                        user = User.query.filter_by(id=user.id).first()
                        user.profile_pic = base64_image
                        db.session.commit()
                return redirect(url_for('user_profile'))
            else:
                return redirect('/login-signup')
        else:
            return redirect('/login-signup')
    return redirect(url_for('user_profile'))




########SPESSSS APPLICANTS/MEMBERS#############
@app.route('/spes_login', methods=['GET', 'POST'])
def spes_login():
    if request.method == 'POST':
        spes_email = request.form['spes_email']
        spes_password = request.form['spes_password']
        spes = SPES.query.filter_by(spes_email=spes_email).first()

        if spes and spes.check_password(spes_password):
            session['spes_email'] = spes.spes_email
            session['spes_application_status'] = spes.spes_application_status  # Assuming this is how you fetch application status
            return redirect('/spes_profile')  # Corrected redirection
        else:
            error = 'Invalid email or password. Please try again.'
            return render_template('spes-login.html', error=error)

    return render_template('spes-login.html')



@app.route('/spes_register', methods=['GET', 'POST'])
def spes_signup():
    if request.method == 'POST':
        spes_name = request.form['spes_name']
        spes_email = request.form['spes_email']
        spes_password = request.form['spes_password']
        confirm_password = request.form['confirm_password']

        existing_user = SPES.query.filter_by(spes_email=spes_email).first()
        if existing_user:
            error = "Email already exists. Please choose a different email."
            return render_template('spes-register.html', error=error)

        if spes_password != confirm_password:
            error = "Passwords do not match!"
            return render_template('spes-register.html', error=error)

        verification_code = ''.join(random.choices('0123456789', k=5))

        send_verification_code_email(spes_email, verification_code)

        session['signup_details'] = {
            'spes_name': spes_name,  # Corrected key name
            'spes_email': spes_email,  # Corrected key name
            'spes_password': spes_password,  # Corrected key name
        }
        session['verification_code'] = verification_code

        flash('Verification code has been sent to your email address. Please check your inbox.', 'success')
        return redirect('/spes_verify')

    return render_template('spes-register.html')

@app.route('/spes_verify', methods=['GET', 'POST'])
def spes_verify():
    if request.method == 'POST':
        entered_code = request.form['verification_code']
        
        if entered_code == session.get('verification_code'):
            signup_details = session.get('signup_details')
            new_spes = SPES(spes_name=signup_details['spes_name'],
                            spes_email=signup_details['spes_email'], 
                            spes_password=signup_details['spes_password'],
                            spes_status='active'  # Set user_status here
                            )
            db.session.add(new_spes)
            db.session.commit()

            session.pop('signup_details')
            session.pop('verification_code')
            flash('Signup successful! You can now log in.', 'success')
            return redirect('/spes_login')
        else:
            error = ('Invalid verification code. Please try again.')
            return render_template('spes-verify.html', error=error)

    return render_template('spes-verify.html')

@app.route('/spes_profile')
def spes_profile():
    if session.get('spes_email'):
        spes = SPES.query.filter_by(spes_email=session['spes_email']).first()
        return render_template('spes-profile.html', spes=spes)
    return redirect('/spes_login')

@app.route('/job_seeker_form_preview')
def job_seeker_form_preview(): 
    if session.get('email'):
        user = User.query.filter_by(email=session['email']).first()
        return render_template('job-applicant-form.html', user=user)
    return redirect('/applicant_login')

@app.route('/edit_profile', methods=['POST'])
def edit_profile():
    if 'email' in session:
        user = User.query.filter_by(email=session['email']).first()
        if user:
            print(request.form)  # Debug statement to print form data

            try:
                # Update user fields
                user.firstname = request.form.get('firstname')
                user.middlename = request.form.get('middlename')
                user.lastname = request.form.get('lastname')
                user.suffix = request.form.get('suffix')
                user.address = request.form.get('address')
                user.gender = request.form.get('gender')
                user.phone = request.form.get('phone')
                user.civil_status = request.form.get('civil_status')
                user.height = request.form.get('height')
                user.landline = request.form.get('landline')
                user.houseNoStreet = request.form.get('houseNoStreet')
                user.brgy = request.form.get('brgy')
                user.city = request.form.get('city')
                user.province = request.form.get('province')
                user.religion = request.form.get('religion')
                user.tinID = request.form.get('tinID')
                user.gsisNO = request.form.get('gsisNO')
                user.pagibigNo = request.form.get('pagibigNo')
                user.philNo = request.form.get('philNo')
                user.disability = request.form.get('disability')
                user.employmentType = request.form.get('employmentType')
                user.employmentStatus = request.form.get('employmentStatus')
                user.activeWork = request.form.get('activeWork')
                user.willingToWork = request.form.get('willingToWork')
                user.porPs = request.form.get('porPs')
                user.prefoccupation = request.form.get('prefoccupation')
                user.preflocation = request.form.get('preflocation')
                user.salary = request.form.get('salary')
                user.passportNo = request.form.get('passportNo')
                user.elemschool = request.form.get('elemschool')
                user.yearGradElem = request.form.get('yearGradElem')
                user.secondarySchool = request.form.get('secondarySchool')
                user.yearGradSec = request.form.get('yearGradSec')
                user.terSchool = request.form.get('terSchool')
                user.courseTer = request.form.get('courseTer')
                user.yearGradTer = request.form.get('yearGradTer')
                user.gradSchool = request.form.get('gradSchool')
                user.courseGrad = request.form.get('courseGrad')
                user.yearGradGrad = request.form.get('yearGradGrad')
                user.skills = request.form.get('skills')
                user.aboutme = request.form.get('aboutme')

                # Handle birthday
                birthday_str = request.form.get('birthday')
                if birthday_str:
                    try:
                        birthday = datetime.strptime(birthday_str, '%m/%d/%Y').date()
                    except ValueError:
                        try:
                            birthday = datetime.strptime(birthday_str, '%Y-%m-%d').date()
                        except ValueError:
                            birthday = None  # Handle invalid date format
                    user.birthday = birthday

                # Handle passport expiry date
                passport_exp_str = request.form.get('passportExp')
                if passport_exp_str:
                    try:
                        passport_exp = datetime.strptime(passport_exp_str, '%m/%d/%Y').date()
                    except ValueError:
                        try:
                            passport_exp = datetime.strptime(passport_exp_str, '%Y-%m-%d').date()
                        except ValueError:
                            passport_exp = None  # Handle invalid date format
                    user.passportExp = passport_exp

                db.session.commit()
                flash('Profile updated successfully', 'success')
                return redirect(url_for('user_profile'))
            except Exception as e:
                db.session.rollback()
                flash(f'Error updating profile: {str(e)}', 'danger')
                print(f'Error: {str(e)}')
                return redirect(url_for('edit_profile'))
        else:
            flash('User not found', 'danger')
            return redirect('/login-signup')
    else:
        flash('You need to log in first', 'warning')
        return redirect('/login-signup')

@app.route('/spes_edit_profile', methods=['POST'])
def spes_edit_profile():
    if 'spes_email' in session:
        spes = SPES.query.filter_by(spes_email=session['spes_email']).first()
        if spes:
            print(request.form) 
            try:

                spes.spes_firstname = request.form.get('spes_firstname')
                spes.spes_middlename = request.form.get('spes_middlename')
                spes.spes_lastname = request.form.get('spes_lastname')
                spes.spes_email = request.form.get('spes_email')
                spes.spes_address = request.form.get('spes_address')
                spes.spes_gender = request.form.get('spes_gender')
                spes.spes_phoneno = request.form.get('spes_phoneno')
                spes.spes_civil = request.form.get('spes_civil')
                spes.spes_spestype = request.form.get('spes_spestype')
                spes.spes_mofirstname = request.form.get('spes_mofirstname')
                spes.spes_momiddlename = request.form.get('spes_momiddlename')
                spes.spes_molastname = request.form.get('spes_molastname')
                spes.spes_fafirstname = request.form.get('spes_fafirstname')
                spes.spes_famiddlename = request.form.get('spes_famiddlename')
                spes.spes_falastname = request.form.get('spes_falastname')
                spes.spes_elemschool = request.form.get('spes_elemschool')
                spes.spes_elemgrad = request.form.get('spes_elemgrad')
                spes.spes_junschool = request.form.get('spes_junschool')
                spes.spes_jungrad = request.form.get('spes_jungrad')
                spes.spes_senschool = request.form.get('spes_senschool')
                spes.spes_sengrad = request.form.get('spes_sengrad')
                spes.spes_strand = request.form.get('spes_strand')
                spes.spes_colschool = request.form.get('spes_colschool')
                spes.spes_course = request.form.get('spes_course')
                spes.spes_colgrad = request.form.get('spes_colgrad')
                spes.spes_facontact = request.form.get('spes_facontact')
                spes.spes_mocontact = request.form.get('spes_mocontact')
                spes.spes_moocupation = request.form.get('spes_moocupation')
                spes.spes_faocupation = request.form.get('spes_faocupation')
                spes.spes_citizenship = request.form.get('spes_citizenship')
                spes.spes_gsisbeneficiary = request.form.get('spes_gsisbeneficiary')
                spes.spes_social = request.form.get('spes_social')
                spes.spes_gsisstatus = request.form.get('spes_gsisstatus')
                spes.spes_presentadd = request.form.get('spes_presentadd')
                spes.spes_permanentadd = request.form.get('spes_permanentadd')
                spes.spes_elemlvl = request.form.get('spes_elemlvl')
                spes.spes_junlvl = request.form.get('spes_junlvl')
                spes.spes_collvl = request.form.get('spes_collvl')

            # Handle birthday field
                spes_birthday_str = request.form.get('spes_birthday')
                if spes_birthday_str:  # Check if the value is not empty or None
                    try:
                        spes_birthday = datetime.strptime(spes_birthday_str, '%Y-%m-%d').date()
                    except ValueError:
                        try:
                            spes_birthday = datetime.strptime(spes_birthday_str, '%m/%d/%Y').date()
                        except ValueError:
                            spes_birthday = None  # Handle the case where the date format is invalid
                else:
                    spes_birthday = None

                spes.spes_birthday = spes_birthday

                db.session.commit()
                return redirect(url_for('spes_profile'))
            except Exception as e:
                    db.session.rollback()
                    flash(f'Error updating profile: {str(e)}', 'danger')
                    print(f'Error: {str(e)}')
                    return redirect(url_for('spes_edit_profile'))
        else:
            flash('User not found', 'danger')
            return redirect('/spes_login')
    else:
        flash('You need to log in first', 'warning')
        return redirect('/spes_login')

@app.route('/upload_documents', methods=['POST'])
def upload_documents():
    if session.get('spes_email'):
        spes = SPES.query.filter_by(spes_email=session['spes_email']).first()

        # Accessing uploaded files from the form
        grades_file = request.files['grades']
        birth_certificate_file = request.files['birth_certificate']
        resume_file = request.files['resume']

        # If files are uploaded, save them to the database
        if grades_file:
            spes.grades = grades_file.read()
        if birth_certificate_file:
            spes.birth_certificate = birth_certificate_file.read()
        if resume_file:
            spes.resume = resume_file.read()

        # Commit changes to the database
        db.session.commit()

    return redirect('/spes_documents')

@app.route('/spes_documents')
def spes_documents():
    if session.get('spes_email'):
        spes = SPES.query.filter_by(spes_email=session['spes_email']).first()
        
        # Fetch the uploaded files from the database
        uploaded_resume = spes.resume
        uploaded_birth_certificate = spes.birth_certificate
        uploaded_grades = spes.grades

        # Print the base64 encoded data for debugging
        # Inside the spes_documents route
        print("Uploaded Resume:", uploaded_resume)
        print("Uploaded Birth Certificate:", uploaded_birth_certificate)
        print("Uploaded Grades:", uploaded_grades)

        
        # Encode the files to base64
        if uploaded_resume:
            uploaded_resume = base64.b64encode(uploaded_resume).decode('utf-8')
        if uploaded_birth_certificate:
            uploaded_birth_certificate = base64.b64encode(uploaded_birth_certificate).decode('utf-8')
        if uploaded_grades:
            uploaded_grades = base64.b64encode(uploaded_grades).decode('utf-8')
        
        return render_template('/spes-viewdocu.html', spes=spes, 
                               uploaded_resume=uploaded_resume, 
                               uploaded_birth_certificate=uploaded_birth_certificate, 
                               uploaded_grades=uploaded_grades)


@app.route('/spes_change_profile', methods=['POST'])
def spes_change_profile():
    if request.method == 'POST':
        if session.get('spes_email'):
            spes = SPES.query.filter_by(spes_email=session['spes_email']).first()
            if spes:
                if 'spes_profile_pic' in request.files:
                    spes_profile_pic = request.files['spes_profile_pic']
                    if spes_profile_pic.filename != '':
                        image_data = spes_profile_pic.read()
                        base64_image = base64.b64encode(image_data).decode('utf-8')
                        spes.spes_profile_pic = base64_image  # Note: Make sure this is the correct attribute
                        db.session.commit()
                return redirect(url_for('spes_profile'))
            else:
                return redirect('/spes_login')
        else:
            return redirect('/spes_login')
    return redirect(url_for('spes_profile'))


@app.route('/upload_documents', methods=['POST'])
def spes_upload_documents():
    if request.method == 'POST' and session.get('spes_email'):
        spesapplicant = SPES.query.filter_by(spes_email=session['spes_email']).first()
        if spesapplicant:
            spes = SPES.query.filter_by(spes_id=spesapplicant.spes_id).first()  # Fetch the PESO associated with the SPES
            if not spes:
                flash('PESO not found for this SPES. Please contact support.', 'error')
                return redirect(url_for('spes_home'))

            resume = request.files['resume']
            birth_certificate = request.files['birth_certificate']
            grades = request.files['grades']

            # Check for existing application
            existing_application = SPESApplication.query.filter_by(spes_id=spesapplicant.spes_id).first()
            if existing_application:
                flash('You have already applied for this job. Please wait for the results of your application process.', 'error')
                return redirect(url_for('spes_home'))

            if resume and birth_certificate and grades:
                # Read file contents
                resume_data = resume.read()
                birth_certificate_data = birth_certificate.read()
                grades_data = grades.read()

                new_application = SPES(
                    spes_id=spesapplicant.spes_id,
                    peso_id=1,  # Assign the peso_id
                    firstname=spesapplicant.spes_firstname,  # Populate fullname from SPES model
                    email=spesapplicant.spes_email,  # Populate email from SPES model
                    address=spesapplicant.spes_address,  # Populate address from SPES model
                    contact=spesapplicant.spes_phoneno,  # Populate contact from SPES model
                    resume=resume_data,
                    birth_certificate=birth_certificate_data,
                    grades=grades_data,
                    spes_application_status='Pending',
                    spes_application_date=datetime.now(),
                )

                #spes.grades = grades_data
                #spes.birth_certificate = birth_certificate_data
               # spes.resume = resume_data
                spes.spes_application_status = 'Pending'
                #spes.spes_application_date=datetime.now(),
                db.session.commit()

                db.session.add(new_application)
                db.session.commit()
                flash('Application submitted successfully!', 'success')
                return redirect(url_for('spes_home'))
            else:
                flash('Please upload all required files!', 'error')
                return redirect(url_for('spes_home'))
        else:
            flash('User not found. Please login.', 'error')
            return redirect(url_for('spes_login'))
    else:
        flash('Please login to submit an application.', 'error')
        return redirect(url_for('spes_login'))


@app.route('/spes_apply_submit', methods=['POST'])
def spes_apply_submit():
    if request.method == 'POST' and session.get('spes_email'):
        spesapplicant = SPES.query.filter_by(spes_email=session['spes_email']).first()
        if spesapplicant:
            spes = SPES.query.filter_by(spes_id=spesapplicant.spes_id).first() 
             # Fetch the PESO associated with the SPES
            if not spes:
                flash('PESO not found for this SPES. Please contact support.', 'error')
                return redirect(url_for('spes_home'))

            resume = spes.resume
            birth_certificate = spes.birth_certificate
            grades = spes.grades

            # Check for existing application
            existing_application = SPESApplication.query.filter_by(spes_id=spesapplicant.spes_id).first()
            if existing_application:
                flash('You have already applied for this job. Please wait for the results of your application process.', 'error')
                return redirect(url_for('spes_home'))

            if resume and birth_certificate and grades:
                # Read file contents

                # Create new SPESApplication object
                new_application = SPESApplication(
                    spes_id=spesapplicant.spes_id,
                    peso_id=1,  # Assign the peso_id
                    fullname=spes.spes_firstname + "‎" + spes.spes_middlename + "‎" +spes.spes_lastname + "‎" ,  # Populate fullname from SPES model
                    email=spesapplicant.spes_email,  # Populate email from SPES model
                    address=spesapplicant.spes_address,  # Populate address from SPES model
                    contact=spesapplicant.spes_phoneno,  # Populate contact from SPES model
                    application_status='Pending',
                    application_date=datetime.now(),
                )


                spes.resume = resume
                spes.birth_certificate = birth_certificate
                spes.grades = grades
                spes.spes_application_status = 'Pending'
                db.session.commit()

                db.session.add(new_application)
                db.session.commit()
                flash('Application submitted successfully!', 'success')
                return redirect(url_for('spes_home'))
            else:
                flash('Please upload all required files!', 'error')
                return redirect(url_for('spes_home'))
        else:
            flash('User not found. Please login.', 'error')
            return redirect(url_for('spes_login'))
    else:
        flash('Please login to submit an application.', 'error')
        return redirect(url_for('spes_login'))




@app.route('/spes_home')
def spes_home():
    return render_template ('/spes-home.html')

@app.route('/spes_apply')
def spes_apply():
    spesapplicant = SPES.query.filter_by(spes_email=session['spes_email']).first()
    return render_template('spes-apply.html', spes=spesapplicant)

@app.route('/spes_applicants')
def spes_applicants():
    application_statuses = ['Pending', 'Under Review', 'Rejected']
    spesapplications = SPES.query.filter(SPES.spes_application_status.in_(application_statuses)).all()
    return render_template ('/spes_applicants.html', spesapplications=spesapplications, base64=base64)

def spes_send_application_status_email(recipient_email, application_id, new_status, fullname):
    subject = 'Application Status Update'
    
    body = f"""<!DOCTYPE html>
<html>
<head>
    <style>
        body {{
            font-family: Arial, sans-serif;
            line-height: 1.6;
            color: #333;
        }}
        .highlight {{
            font-size: 15px;
            font-weight: bold;
        }}
    </style>
</head>
<body>
    <p>Dear {fullname},</p>
    <p>We hope this message finds you well. We are writing to inform you that the status of your application for SPES has been updated:</p>
    <p class="highlight">{new_status}</p>
    <p>If you have any questions or need further assistance, please do not hesitate to contact us.</p>
    <p>Thank you for your continued interest and patience.</p>
    <br>
    <p>Best regards,</p>
    <p>PilaHunt Team</p>
</body>
</html>"""

    message = Message(subject, recipients=[recipient_email])
    message.html = body  # Use the 'html' attribute to send HTML content
    mail.send(message)


def employer_send_application_status_email(recipient_email, employerId, new_status, contact_person):
    try:
        subject = f'Application Status Update - Company ID: {employerId}'
        body = f"""<html>
                    <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333; ">
                        <p>Dear {contact_person},</p>
                        <p>We are reaching out to inform you that the status of your application for Employer (Employer ID: {employerId}) has been updated to:</p>
                        <p style="font-size: 15px; font-weight: bold;">{new_status}</p>
                        <p>Best regards,</p>
                        <p>PilaHunt</p>
                    </body>
                   </html>"""

        message = Message(subject, recipients=[recipient_email])
        message.html = body
        mail.send(message)
    except Exception as e:
        print(f"Error sending email: {e}")


def applicant_send_application_status_email(recipient_email, applicantId, new_status, firstname):
    try:
        subject = f'Application Status Update'
        body = f"""<html>
                    <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333; ">
                        <p>Dear {firstname},</p>
                        <p>We are reaching out to inform you that the status of your application for Public Employment Service has been updated to:</p>
                        <p style="font-size: 15px; font-weight: bold;">{new_status}</p>
                        <p>Best regards,</p>
                        <p>PilaHunt</p>
                    </body>
                   </html>"""

        message = Message(subject, recipients=[recipient_email])
        message.html = body
        mail.send(message)
    except Exception as e:
        print(f"Error sending email: {e}")

@app.route('/update_status', methods=['POST'])
def update_status():
    data = request.json
    application_id = data.get('applicationId')
    status = data.get('status')

    if application_id is None:
        return jsonify({'message': 'Application ID is missing'}), 400

    application = SPESApplication.query.get(application_id)
    if application is None:
        return jsonify({'message': 'Application not found'}), 404

    try:
        application.application_status = status
        db.session.commit()

        # Update the corresponding SPES status regardless of the value of 'tatus'
        spes = SPES.query.filter_by(spes_email=application.email).first()
        if spes:
            spes.spes_application_status = status
            db.session.commit()

        # Send email notification
        spes_send_application_status_email(application.email, application_id, status, application.fullname)

        return jsonify({'message': 'Status updated successfully'}), 200
    except Exception as e:
        return jsonify({'message': 'Error updating status'}), 500


@app.route('/employer_update_status', methods=['POST'])
def employer_update_status():
    data = request.json
    employerId = data.get('employerId')
    status = data.get('status')

    # Check if employerId and status are present
    if employerId is None or status is None:
        return jsonify({'error': 'employerId or status missing'}), 400

    # Assuming EmployerDetails is your SQLAlchemy model
    employer = EmployerDetails.query.get(employerId)
    if employer is None:
        return jsonify({'error': 'Employer not found'}), 404

    # Update employer status
    employer.verification_status = status
    db.session.commit()

    # Send email notification
    employer_send_application_status_email(employer.email_add, employerId, status, employer.contact_person)

    return jsonify({'message': 'Status updated successfully'}), 200

@app.route('/applicant_update_status', methods=['POST'])
def applicant_update_status():
    data = request.json
    applicantId = data.get('applicantId')
    status = data.get('status')

    # Check if applicantId and status are present
    if applicantId is None or status is None:
        return jsonify({'error': 'applicantId or status missing'}), 400

    try:
        applicantId = int(applicantId)
    except ValueError:
        return jsonify({'error': 'Invalid applicantId'}), 400

    # Find the user with the given applicantId
    user = User.query.get(applicantId)
    if user is None:
        return jsonify({'error': 'Applicant not found'}), 404

    # Update the user's work status
    user.workStatus = status
    db.session.commit()

    # Send email notification
    applicant_send_application_status_email(user.email, applicantId, status, user.firstname)

    return jsonify({'message': 'Status updated successfully'}), 200

@app.template_filter('dateformat')
def dateformat(value, format='%Y-%m-%d'):
    if value:
        return value.strftime(format)
    return value

# Ensure you import the necessary modules for your app configuration
from flask import session, redirect, url_for, render_template

@app.route('/view_tasks')
def view_tasks():
    spes_email = session.get('spes_email')
    print(f"User email from session: {spes_email}")  # Debug statement
    
    if spes_email:
        spes_user = SPES.query.filter_by(spes_email=spes_email).first()
        print(f"SPES User: {spes_user}")  # Debug statement
        
        if spes_user:
            tasks = Task.query.filter_by(spes_id=spes_user.spes_id).all()
            print(f"Tasks: {tasks}")  # Debug statement
            task_events = [
                {
                    'id': task.task_id,  # Ensure the task ID is included
                    'title': 'Completed' if task.task_status == 'Completed' else task.task_title,
                    'description': task.task_description,
                    'start': task.task_due_date.strftime('%Y-%m-%d') if task.task_status == 'Completed' else task.task_due_date.strftime('%Y-%m-%d %H:%M:%S'),  # Format due date and time
                    'backgroundColor': 'green' if task.task_status == 'Completed' else '#FCE883',  # Set background color based on status
                    'borderColor': 'green' if task.task_status == 'Completed' else '#FCE883',  # Set border color based on status
                    'status': task.task_status,
                    'due_date': task.task_due_date.strftime('%Y-%m-%d %I:%M %p') if task.task_due_date else 'No due date'
                } for task in tasks
            ]
            pending_tasks = Task.query.filter_by(spes_id=spes_user.spes_id, task_status='Pending').all()
            pending_count = len(pending_tasks)
            formatted_due_dates = [task.task_due_date.strftime('%A, %B %d, %I:%M %p').lstrip("0").replace(" 0", " ") for task in pending_tasks if task.task_due_date]

        else:
            task_events = []
            pending_count = 0
            formatted_due_dates = []
            print("No SPES user found.")  # Debug statement
    else:
        print("No user email found in session. Redirecting to login.")  # Debug statement
        return redirect(url_for('spes_login'))

    print(f"Task Events: {task_events}")  # Debug statement
    print(f"Pending Count: {pending_count}")  # Debug statement
    print(f"Formatted Due Dates: {formatted_due_dates}")  # Debug statement

    return render_template('view_tasks.html', tasks=task_events, pending_count=pending_count, formatted_due_dates=formatted_due_dates)


@app.route('/complete_task', methods=['POST'])
def complete_task():
    if 'taskFile' not in request.files or 'taskId' not in request.form:
        return jsonify({'error': 'Bad request, missing parameters'}), 400

    task_file = request.files['taskFile']
    task_id = request.form['taskId']

    if task_file.filename == '':
        return jsonify({'error': 'No selected file'}), 400

    task = Task.query.get(task_id)
    if not task:
        return jsonify({'error': 'Task not found'}), 404

    # Read the file and save it as a binary data
    task.file = task_file.read()
    task.task_status = 'Completed' 

    db.session.commit()
    return jsonify({'success': 'Task updated successfully'}), 200


@app.route('/date_time_picker')
def date_time_Picker():
    return render_template ('/date_time_picker.html')


##########PESO#########################


@app.route('/spes_members')
def spes_members():
    approved_applicants = SPESApplication.query.filter_by(application_status='Approved').all()
    return render_template('spes_members.html', applicants=approved_applicants)


@app.route('/save_task', methods=['POST'])
def save_task():
    data = request.json
    application_id = data['applicationId']
    task_title = data['taskTitle']
    task_description = data['taskDescription']
    
    # Convert the task_due_date string to a Python datetime object
    task_due_date_str = data['taskDueDate']
    task_due_date = datetime.strptime(task_due_date_str, '%Y-%m-%dT%H:%M')

    # Assuming you have imported the Task model from your models.py
    task = Task(
        spes_id = application_id,
        task_title=task_title,
        task_description=task_description,
        peso_id = "1",
        task_due_date=task_due_date,
        task_status='Pending',
        # Assuming you have the required fields here
    )
    db.session.add(task)
    db.session.commit()

    return jsonify({'message': 'Task saved successfully'})

@app.route('/resume_preview')
def resume_preview():
    if session.get('spes_email'):
        spes = SPES.query.filter_by(spes_email=session['spes_email']).first()
        return render_template('resume.html', spes=spes)
    return render_template ('/spes_login')

def newline_every_seven_words(value):
    if not value:
        return ''
    words = value.split()
    lines = []
    for i in range(0, len(words), 7):
        lines.append(' '.join(words[i:i+7]))
    return '<br>'.join(lines)

def skills_to_bullets(skills):
    if not skills:
        return ''
    skill_list = skills.split(',')
    bullet_points = ''.join(f'<li>{skill.strip()}</li>' for skill in skill_list)
    return f'<ul>{bullet_points}</ul>'

app.jinja_env.filters['newline_every_seven_words'] = newline_every_seven_words
app.jinja_env.filters['skills_to_bullets'] = skills_to_bullets

@app.route('/resume_preview_applicant')
def resume_preview_applicant():
    try:
        if session.get('email'):
            user = User.query.filter_by(email=session['email']).first()
            if user:
                print(user.aboutme)  # Debugging line
                return render_template('resume-applicant.html', user=user)
        return render_template('applicant_login.html')
    except Exception as e:
        print(f"Error: {e}")  # Debugging line
        return render_template('applicant_login.html')

    
@app.route('/peso_login')
def peso_login():
    return render_template ('/admin-login.html')

@app.route('/spes_form')
def spes_form():
    if session.get('spes_email'):
        spes = SPES.query.filter_by(spes_email=session['spes_email']).first()
        return render_template('spes-form.html', spes=spes)
    return redirect('/spes_login')

@app.route('/peso_dashboard')
def peso_dashboard():
    from sqlalchemy import desc, or_

    applicant_count = User.query.filter_by(user_type='applicant').count()
    employer_count = User.query.filter_by(user_type='employer').count()
    total_count = User.query.filter(or_(User.user_type=='applicant', User.user_type=='employer')).count()

    # Order by primary key in descending order to get the newest entries first
    applicants = User.query.filter_by(user_type='applicant').order_by(desc(User.id)).all()
    employers = EmployerDetails.query.filter_by(verification_status='Approved').order_by(desc(EmployerDetails.company_id)).all()
    jobs = Jobs.query.order_by(desc(Jobs.id)).all()
    job_count = Jobs.query.count()

    print([applicant.firstname for applicant in applicants])  # Debug: Print out first names

    return render_template('peso_dashboard.html', 
                           applicant_count=applicant_count, 
                           employer_count=employer_count, 
                           total_count=total_count, 
                           applicants=applicants, 
                           employers=employers, 
                           jobs=jobs, 
                           job_count=job_count)  # Fix variable name

@app.route('/admin_forgot_pass')
def admin_forgot_pass():
    return render_template ('/admin_forgot_pass.html')


@app.route('/employer_list')
def employer_list():
    employers = EmployerDetails.query.filter_by(verification_status='Approved').all()
    return render_template('employer_list.html', employers=employers)

@app.route('/approval_request', methods=['GET', 'POST'])
def approval_request():
    if request.method == 'GET':
        waiting_verification = EmployerDetails.query.filter(
    or_(
        EmployerDetails.verification_status == "waiting verification",
        EmployerDetails.verification_status == "Under Review",
        EmployerDetails.verification_status == "not verified"
    )
).all()
        return render_template('approval_request.html', waiting_verification=waiting_verification)
    elif request.method == 'POST':

        company_id = request.form['company_id']
        verification_status = request.form['verification_status']
        employer = EmployerDetails.query.get(company_id)
        if employer:
            employer.verification_status = verification_status
            db.session.commit()
        return redirect(url_for('approval_request'))


@app.route('/rejected_request', methods=['GET', 'POST'])
def rejected_request():
    if request.method == 'GET':
        waiting_verification = EmployerDetails.query.filter(
    or_(
        EmployerDetails.verification_status == "Denied Request",
    )
).all()
        return render_template('rejected_request.html', waiting_verification=waiting_verification)
    elif request.method == 'POST':

        company_id = request.form['company_id']
        verification_status = request.form['verification_status']
        employer = EmployerDetails.query.get(company_id)
        if employer:
            employer.verification_status = verification_status
            db.session.commit()
        return redirect(url_for('rejected_request'))

@app.route('/applicant_lists')
def applicant_lists():
    applicants = User.query.filter_by(user_type='applicant').all()
    return render_template('applicants.html', applicants=applicants)

@app.route('/export_applicants/pdf')
def export_applicants_to_pdf():
    applicants = User.query.filter_by(user_type='applicant').all()
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    y = height - 40
    p.drawString(100, y, "Full Name | Email | Address | Contact | Gender")

    for applicant in applicants:
        y -= 20
        address = f"{applicant.houseNoStreet} {applicant.brgy} {applicant.city}"
        p.drawString(100, y, f"{applicant.firstname} {applicant.lastname} | {applicant.email} | {address} | 0{applicant.phone} | {applicant.gender}")

    p.showPage()
    p.save()
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="applicants.pdf", mimetype='application/pdf')

@app.route('/export_applicants/excel')
def export_applicants_to_excel():
    applicants = User.query.filter_by(user_type='applicant').all()
    buffer = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applicants"

    columns = ['Full Name', 'Email', 'Address', 'Contact', 'Gender']
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_title

    for applicant in applicants:
        row_num += 1
        full_name = f"{applicant.firstname} {applicant.lastname}"
        address = f"{applicant.houseNoStreet} {applicant.brgy} {applicant.city}"
        ws.cell(row=row_num, column=1).value = full_name
        ws.cell(row=row_num, column=2).value = applicant.email
        ws.cell(row=row_num, column=3).value = address
        ws.cell(row=row_num, column=4).value = f"0{applicant.phone}"
        ws.cell(row=row_num, column=5).value = applicant.gender

    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="applicants.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/export/pdf')
def export_employers_to_pdf():
    employers = EmployerDetails.query.filter_by(verification_status='Approved').all()
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    y = height - 40
    p.drawString(100, y, "Company Name | TIN ID | Address | Contact Person | Phone | Business Type | Status")

    for employer in employers:
        y -= 20
        p.drawString(100, y, f"{employer.company_name} | {employer.tin_id} | {employer.company_address} | {employer.contact_person} | {employer.phone_no} | {employer.business_type} | {employer.verification_status}")

    p.showPage()
    p.save()
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="employers.pdf", mimetype='application/pdf')

@app.route('/export/excel')
def export_employers_to_excel():
    employers = EmployerDetails.query.filter_by(verification_status='Approved').all()
    buffer = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employers"

    columns = ['Company Name', 'TIN ID', 'Address', 'Contact Person', 'Phone', 'Business Type', 'Status']
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_title

    for employer in employers:
        row_num += 1
        ws.cell(row=row_num, column=1).value = employer.company_name
        ws.cell(row=row_num, column=2).value = employer.tin_id
        ws.cell(row=row_num, column=3).value = employer.company_address
        ws.cell(row=row_num, column=4).value = employer.contact_person
        ws.cell(row=row_num, column=5).value = employer.phone_no
        ws.cell(row=row_num, column=6).value = employer.business_type
        ws.cell(row=row_num, column=7).value = employer.verification_status

    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="employers.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


if __name__ == '__main__':
    app.run(debug=True)
